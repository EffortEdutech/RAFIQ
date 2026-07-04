#!/usr/bin/env python3
"""Run RAFIQ Day 8 representative imports and executable validation rules."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import sqlite3
import sys
import uuid
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT = ROOT / "data" / "staging_reports" / "day8"
RUN_ID = "day8-representative-import-v1"
NAMESPACE = uuid.UUID("391dbcec-bb30-4aa9-b61c-5607d2c3386a")
AYAH_COUNTS = (
    7, 286, 200, 176, 120, 165, 206, 75, 129, 109, 123, 111, 43, 52, 99,
    128, 111, 110, 98, 135, 112, 78, 118, 64, 77, 227, 93, 88, 69, 60, 34,
    30, 73, 54, 45, 83, 182, 88, 75, 85, 54, 53, 89, 59, 37, 35, 38, 29,
    18, 45, 60, 49, 62, 55, 78, 96, 29, 22, 24, 13, 14, 11, 11, 18, 12,
    12, 30, 52, 52, 44, 28, 28, 20, 56, 40, 31, 50, 40, 46, 42, 29, 19,
    36, 25, 22, 17, 19, 26, 30, 20, 15, 21, 11, 8, 8, 19, 5, 8, 8, 11,
    11, 8, 3, 9, 5, 4, 7, 3, 6, 3, 5, 4, 5, 6,
)

SOURCES = {
    "quran": ("qul-quran-script-uthmani-88", ROOT / "data/raw/quran/qul/uthmani.json"),
    "translation": (
        "qul-translation-saheeh-193",
        ROOT / "data/raw/translations/qul/en-sahih-international-simple.json",
    ),
    "tafsir": (
        "qul-tafsir-english-mukhtasar-266",
        ROOT / "data/raw/tafsir/en-tafisr-Al-Mukhtasar.json",
    ),
    "topics": ("qul-topics-45", ROOT / "data/raw/tafsir/topics.db"),
    "themes": ("qul-ayah-themes-62", ROOT / "data/raw/tafsir/ayah-themes.db"),
    "hadith": (
        "fawaz-hadith-api-eng-abudawud",
        ROOT / "data/raw/hadith/collections/fawaz-hadith-api-v1/editions/eng-abudawud.json",
    ),
    "verification": (
        "semakhadis-api-hadith-seeder",
        ROOT / "data/raw/hadith/verification/semakhadis-api/database/seeds/seeder_csv/hadith_seeder.xlsx",
    ),
}


def utcnow() -> str:
    return datetime.now(timezone.utc).isoformat()


def stable_id(*parts: object) -> str:
    return str(uuid.uuid5(NAMESPACE, "|".join(str(part) for part in parts)))


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def json_load(path: Path):
    return json.loads(path.read_text(encoding="utf-8"))


def ayah_key_valid(key: str) -> bool:
    match = re.fullmatch(r"(\d+):(\d+)", str(key))
    if not match:
        return False
    surah, ayah = map(int, match.groups())
    return 1 <= surah <= 114 and 1 <= ayah <= AYAH_COUNTS[surah - 1]


def split_csv_values(value: object) -> list[str]:
    if value is None:
        return []
    return [part.strip() for part in str(value).split(",") if part.strip()]


class Prototype:
    def __init__(self, output_dir: Path):
        self.output_dir = output_dir
        self.db_path = output_dir / "import_prototype.sqlite"
        self.results: list[dict[str, object]] = []
        self.metrics: dict[str, object] = {}
        self.conn: sqlite3.Connection | None = None

    def finding(
        self,
        domain: str,
        code: str,
        severity: str,
        passed: bool,
        observed: object,
        expected: object,
        notes: str = "",
        source_key: str | None = None,
    ) -> None:
        row = {
            "domain": domain,
            "rule_code": code,
            "severity": severity,
            "result": "PASS" if passed else "FAIL",
            "observed": observed,
            "expected": expected,
            "source_record_key": source_key or "",
            "notes": notes,
        }
        self.results.append(row)
        self.conn.execute(
            """INSERT INTO validation_findings
               (finding_id, import_run_id, domain, rule_code, severity, result,
                observed_value, expected_value, source_record_key, notes)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                stable_id(RUN_ID, domain, code, source_key or "", len(self.results)),
                RUN_ID,
                domain,
                code,
                severity,
                row["result"],
                str(observed),
                str(expected),
                source_key,
                notes,
            ),
        )

    def source_record(
        self, source_id: str, key: str, path: Path, locator: str, raw: object
    ) -> str:
        record_id = stable_id(source_id, key)
        payload = json.dumps(raw, ensure_ascii=False, sort_keys=True, default=str)
        self.conn.execute(
            """INSERT INTO source_records
               (source_record_id, import_run_id, source_id, source_record_key,
                raw_object_path, source_locator, raw_record_json, raw_record_sha256)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                record_id,
                RUN_ID,
                source_id,
                key,
                path.relative_to(ROOT).as_posix(),
                locator,
                payload,
                hashlib.sha256(payload.encode("utf-8")).hexdigest(),
            ),
        )
        return record_id

    def create_schema(self) -> None:
        self.conn.executescript(
            """
            PRAGMA foreign_keys = ON;
            CREATE TABLE import_runs (
              import_run_id TEXT PRIMARY KEY, parser_name TEXT NOT NULL,
              parser_version TEXT NOT NULL, started_at TEXT NOT NULL,
              completed_at TEXT, status TEXT NOT NULL
            );
            CREATE TABLE raw_objects (
              source_id TEXT NOT NULL, path TEXT NOT NULL, byte_length INTEGER NOT NULL,
              sha256 TEXT NOT NULL, PRIMARY KEY (source_id, path)
            );
            CREATE TABLE source_records (
              source_record_id TEXT PRIMARY KEY, import_run_id TEXT NOT NULL,
              source_id TEXT NOT NULL, source_record_key TEXT NOT NULL,
              raw_object_path TEXT NOT NULL, source_locator TEXT NOT NULL,
              raw_record_json TEXT NOT NULL, raw_record_sha256 TEXT NOT NULL,
              UNIQUE (source_id, source_record_key)
            );
            CREATE TABLE validation_findings (
              finding_id TEXT PRIMARY KEY, import_run_id TEXT NOT NULL,
              domain TEXT NOT NULL, rule_code TEXT NOT NULL, severity TEXT NOT NULL,
              result TEXT NOT NULL, observed_value TEXT, expected_value TEXT,
              source_record_key TEXT, notes TEXT
            );
            CREATE TABLE quran_ayah_texts (
              record_id TEXT PRIMARY KEY, source_record_id TEXT NOT NULL,
              ayah_key TEXT NOT NULL UNIQUE, surah INTEGER NOT NULL, ayah INTEGER NOT NULL,
              script_variant TEXT NOT NULL, text TEXT NOT NULL
            );
            CREATE TABLE translation_texts (
              record_id TEXT PRIMARY KEY, source_record_id TEXT NOT NULL,
              ayah_key TEXT NOT NULL, variant TEXT NOT NULL, language TEXT NOT NULL,
              text TEXT, structured_json TEXT, UNIQUE (ayah_key, variant)
            );
            CREATE TABLE translation_footnotes (
              record_id TEXT PRIMARY KEY, translation_record_id TEXT NOT NULL,
              footnote_key TEXT NOT NULL, footnote_text TEXT NOT NULL
            );
            CREATE TABLE tafsir_passages (
              passage_id TEXT PRIMARY KEY, source_record_id TEXT NOT NULL,
              source_key TEXT NOT NULL UNIQUE, text TEXT NOT NULL
            );
            CREATE TABLE tafsir_passage_ayahs (
              passage_id TEXT NOT NULL, ayah_key TEXT NOT NULL,
              relation_type TEXT NOT NULL, source_record_key TEXT NOT NULL,
              PRIMARY KEY (passage_id, ayah_key)
            );
            CREATE TABLE source_topics (
              record_id TEXT PRIMARY KEY, source_record_id TEXT NOT NULL,
              topic_id INTEGER NOT NULL UNIQUE, name TEXT, arabic_name TEXT,
              description TEXT, wiki_link TEXT, thematic INTEGER, ontology INTEGER
            );
            CREATE TABLE source_topic_relations (
              topic_id INTEGER NOT NULL, related_topic_id INTEGER NOT NULL,
              relation_type TEXT NOT NULL,
              PRIMARY KEY (topic_id, related_topic_id, relation_type)
            );
            CREATE TABLE source_topic_ayahs (
              topic_id INTEGER NOT NULL, ayah_key TEXT NOT NULL,
              PRIMARY KEY (topic_id, ayah_key)
            );
            CREATE TABLE ayah_theme_groups (
              record_id TEXT PRIMARY KEY, source_record_id TEXT NOT NULL,
              source_row INTEGER NOT NULL UNIQUE, theme TEXT NOT NULL,
              surah INTEGER NOT NULL, ayah_from INTEGER NOT NULL, ayah_to INTEGER NOT NULL,
              keywords TEXT, total_ayahs INTEGER NOT NULL, exact_group_key TEXT NOT NULL
            );
            CREATE TABLE ayah_theme_group_ayahs (
              theme_record_id TEXT NOT NULL, ayah_key TEXT NOT NULL,
              PRIMARY KEY (theme_record_id, ayah_key)
            );
            CREATE TABLE hadith_records (
              record_id TEXT PRIMARY KEY, source_record_id TEXT NOT NULL,
              collection_key TEXT NOT NULL, hadith_number INTEGER NOT NULL,
              arabic_number INTEGER, book_number INTEGER, book_hadith_number INTEGER,
              section_name TEXT, UNIQUE (collection_key, hadith_number)
            );
            CREATE TABLE hadith_text_versions (
              record_id TEXT PRIMARY KEY, hadith_record_id TEXT NOT NULL,
              language TEXT NOT NULL, text TEXT NOT NULL
            );
            CREATE TABLE hadith_grade_assertions (
              record_id TEXT PRIMARY KEY, hadith_record_id TEXT NOT NULL,
              grader TEXT NOT NULL, raw_grade TEXT NOT NULL
            );
            CREATE TABLE hadith_verification_claims (
              record_id TEXT PRIMARY KEY, source_record_id TEXT NOT NULL,
              source_row INTEGER NOT NULL, arabic_text TEXT, malay_text TEXT,
              companion TEXT, chapter TEXT, keywords TEXT, raw_status TEXT,
              commentary TEXT, researcher TEXT
            );
            CREATE TABLE hadith_verification_references (
              record_id TEXT PRIMARY KEY, claim_id TEXT NOT NULL,
              ordinal INTEGER NOT NULL, reference_text TEXT,
              volume_page_text TEXT
            );
            """
        )
        self.conn.execute(
            "INSERT INTO import_runs VALUES (?, ?, ?, ?, NULL, ?)",
            (RUN_ID, "run_day8_import_prototype.py", "1.0.0", utcnow(), "running"),
        )
        for source_id, path in SOURCES.values():
            self.conn.execute(
                "INSERT INTO raw_objects VALUES (?, ?, ?, ?)",
                (source_id, path.relative_to(ROOT).as_posix(), path.stat().st_size, sha256(path)),
            )

    def load_quran(self) -> set[str]:
        source_id, path = SOURCES["quran"]
        data = json_load(path)
        keys = set()
        blanks = 0
        for key, raw in data.items():
            source_record_id = self.source_record(source_id, key, path, f"/{key}", raw)
            text = str(raw.get("text", ""))
            blanks += not text.strip()
            keys.add(key)
            self.conn.execute(
                "INSERT INTO quran_ayah_texts VALUES (?, ?, ?, ?, ?, ?, ?)",
                (
                    stable_id(source_id, key, "uthmani"),
                    source_record_id,
                    key,
                    raw["surah"],
                    raw["ayah"],
                    "uthmani",
                    text,
                ),
            )
        self.finding("quran", "QURAN_RECORD_COUNT", "error", len(data) == 6236, len(data), 6236)
        self.finding("quran", "QURAN_UNIQUE_KEYS", "error", len(keys) == len(data), len(keys), len(data))
        invalid = sum(not ayah_key_valid(key) for key in keys)
        self.finding("quran", "QURAN_KEY_RANGE", "error", invalid == 0, invalid, 0)
        self.finding("quran", "QURAN_BLANK_TEXT", "error", blanks == 0, blanks, 0)
        self.metrics["quran"] = {"records": len(data), "blank_text": blanks}
        return keys

    def load_translations(self, canonical_keys: set[str]) -> None:
        source_id, simple_path = SOURCES["translation"]
        variants = {
            "simple": simple_path,
            "inline_footnotes": simple_path.with_name("en-sahih-international-inline-footnotes.json"),
            "footnote_tags": simple_path.with_name("en-sahih-international-with-footnote-tags.json"),
            "chunks": simple_path.with_name("en-sahih-international-chunks.json"),
        }
        counts = {}
        missing_references = 0
        footnote_count = 0
        for variant, path in variants.items():
            data = json_load(path)
            counts[variant] = len(data)
            for key, raw in data.items():
                source_record_id = self.source_record(
                    f"{source_id}-{variant}", key, path, f"/{key}", raw
                )
                text = raw.get("t")
                structured = None
                if isinstance(text, list):
                    structured = json.dumps(text, ensure_ascii=False)
                    display_text = "".join(part for part in text if isinstance(part, str))
                    references = [str(part["f"]) for part in text if isinstance(part, dict) and "f" in part]
                    missing_references += sum(ref not in raw.get("f", {}) for ref in references)
                else:
                    display_text = str(text or "")
                record_id = stable_id(source_id, variant, key)
                self.conn.execute(
                    "INSERT INTO translation_texts VALUES (?, ?, ?, ?, ?, ?, ?)",
                    (record_id, source_record_id, key, variant, "en", display_text, structured),
                )
                for footnote_key, footnote_text in raw.get("f", {}).items():
                    footnote_count += 1
                    self.conn.execute(
                        "INSERT INTO translation_footnotes VALUES (?, ?, ?, ?)",
                        (
                            stable_id(record_id, footnote_key),
                            record_id,
                            str(footnote_key),
                            str(footnote_text),
                        ),
                    )
            self.finding(
                "translation",
                f"TRANSLATION_{variant.upper()}_COUNT",
                "error",
                len(data) == 6236,
                len(data),
                6236,
            )
            self.finding(
                "translation",
                f"TRANSLATION_{variant.upper()}_KEY_COVERAGE",
                "error",
                set(data) == canonical_keys,
                len(set(data) & canonical_keys),
                len(canonical_keys),
            )
            blanks = sum(not str(item.get("t", "")).strip() for item in data.values())
            self.finding(
                "translation",
                f"TRANSLATION_{variant.upper()}_BLANK_TEXT",
                "error",
                blanks == 0,
                blanks,
                0,
            )
        self.finding(
            "translation",
            "TRANSLATION_CHUNK_FOOTNOTE_REFERENCES",
            "error",
            missing_references == 0,
            missing_references,
            0,
        )
        self.metrics["translation"] = {
            "variant_counts": counts,
            "staged_texts": sum(counts.values()),
            "staged_footnotes": footnote_count,
        }

    def load_tafsir(self, canonical_keys: set[str]) -> None:
        source_id, path = SOURCES["tafsir"]
        data = json_load(path)
        pointers = {}
        passages = {}
        covered = set()
        blank = 0
        for key, raw in data.items():
            source_record_id = self.source_record(source_id, key, path, f"/{key}", raw)
            if isinstance(raw, str):
                pointers[key] = raw
                continue
            text = str(raw.get("text", ""))
            blank += not text.strip()
            passage_id = stable_id(source_id, key, "passage")
            passages[key] = passage_id
            self.conn.execute(
                "INSERT INTO tafsir_passages VALUES (?, ?, ?, ?)",
                (passage_id, source_record_id, key, text),
            )
            ayah_keys = raw.get("ayah_keys") or [key]
            for ayah_key in ayah_keys:
                covered.add(ayah_key)
                self.conn.execute(
                    "INSERT INTO tafsir_passage_ayahs VALUES (?, ?, ?, ?)",
                    (passage_id, ayah_key, "explicit_group" if len(ayah_keys) > 1 else "direct", key),
                )
        invalid_pointer = 0
        for key, target in pointers.items():
            if target not in passages:
                invalid_pointer += 1
                continue
            covered.add(key)
            self.conn.execute(
                "INSERT OR IGNORE INTO tafsir_passage_ayahs VALUES (?, ?, ?, ?)",
                (passages[target], key, "pointer", key),
            )
        self.finding("tafsir", "TAFSIR_SOURCE_RECORD_COUNT", "error", len(data) == 6236, len(data), 6236)
        self.finding("tafsir", "TAFSIR_PASSAGE_COUNT", "error", len(passages) == 6216, len(passages), 6216)
        self.finding("tafsir", "TAFSIR_POINTER_COUNT", "error", len(pointers) == 20, len(pointers), 20)
        self.finding("tafsir", "TAFSIR_POINTER_TARGETS", "error", invalid_pointer == 0, invalid_pointer, 0)
        self.finding(
            "tafsir", "TAFSIR_AYAH_COVERAGE", "error", covered == canonical_keys, len(covered), len(canonical_keys)
        )
        self.finding("tafsir", "TAFSIR_BLANK_PASSAGES", "warning", blank == 0, blank, 0)
        self.metrics["tafsir"] = {
            "source_records": len(data),
            "passages": len(passages),
            "pointers": len(pointers),
            "covered_ayahs": len(covered),
        }

    def load_topics(self, canonical_keys: set[str]) -> None:
        source_id, path = SOURCES["topics"]
        source = sqlite3.connect(f"file:{path.as_posix()}?mode=ro", uri=True)
        source.row_factory = sqlite3.Row
        rows = list(source.execute("SELECT * FROM topics ORDER BY topic_id"))
        source.close()
        topic_ids = {row["topic_id"] for row in rows}
        invalid_relations = invalid_ayahs = ayah_links = relation_count = 0
        for row_number, row in enumerate(rows, 1):
            raw = dict(row)
            key = str(row["topic_id"])
            source_record_id = self.source_record(source_id, key, path, f"topics.topic_id={key}", raw)
            self.conn.execute(
                "INSERT INTO source_topics VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (
                    stable_id(source_id, key),
                    source_record_id,
                    row["topic_id"],
                    row["name"],
                    row["arabic_name"],
                    row["description"],
                    row["wiki_link"],
                    row["thematic"],
                    row["ontology"],
                ),
            )
            relations = [
                ("parent", row["parent_id"]),
                ("thematic_parent", row["thematic_parent_id"]),
                ("ontology_parent", row["ontology_parent_id"]),
            ]
            relations += [("related", value) for value in split_csv_values(row["related_topics"])]
            for relation_type, target in relations:
                if target in (None, "", 0, "0"):
                    continue
                try:
                    target_id = int(target)
                except (TypeError, ValueError):
                    invalid_relations += 1
                    continue
                if target_id not in topic_ids:
                    invalid_relations += 1
                relation_count += 1
                self.conn.execute(
                    "INSERT OR IGNORE INTO source_topic_relations VALUES (?, ?, ?)",
                    (row["topic_id"], target_id, relation_type),
                )
            for ayah_key in split_csv_values(row["ayahs"]):
                ayah_links += 1
                invalid_ayahs += ayah_key not in canonical_keys
                self.conn.execute(
                    "INSERT OR IGNORE INTO source_topic_ayahs VALUES (?, ?)",
                    (row["topic_id"], ayah_key),
                )
        self.finding("topics", "TOPIC_RECORD_COUNT", "error", len(rows) == 2512, len(rows), 2512)
        self.finding("topics", "TOPIC_UNIQUE_IDS", "error", len(topic_ids) == len(rows), len(topic_ids), len(rows))
        self.finding("topics", "TOPIC_RELATION_REFERENCES", "error", invalid_relations == 0, invalid_relations, 0)
        self.finding("topics", "TOPIC_AYAH_REFERENCES", "error", invalid_ayahs == 0, invalid_ayahs, 0)
        self.metrics["topics"] = {
            "records": len(rows),
            "relations": relation_count,
            "ayah_links": ayah_links,
        }

    def load_themes(self, canonical_keys: set[str]) -> None:
        source_id, path = SOURCES["themes"]
        source = sqlite3.connect(f"file:{path.as_posix()}?mode=ro", uri=True)
        source.row_factory = sqlite3.Row
        rows = list(source.execute("SELECT rowid AS source_row, * FROM themes ORDER BY rowid"))
        source.close()
        exact_keys = []
        covered = set()
        invalid_ranges = total_mismatches = 0
        for row in rows:
            raw = dict(row)
            source_row = row["source_row"]
            source_record_id = self.source_record(
                source_id, str(source_row), path, f"themes.rowid={source_row}", raw
            )
            exact = json.dumps(
                [
                    row["theme"], row["surah_number"], row["ayah_from"],
                    row["ayah_to"], row["keywords"], row["total_ayahs"],
                ],
                ensure_ascii=False,
            )
            exact_keys.append(exact)
            valid = (
                1 <= row["surah_number"] <= 114
                and 1 <= row["ayah_from"] <= row["ayah_to"] <= AYAH_COUNTS[row["surah_number"] - 1]
            )
            invalid_ranges += not valid
            expected_total = row["ayah_to"] - row["ayah_from"] + 1
            total_mismatches += row["total_ayahs"] != expected_total
            record_id = stable_id(source_id, source_row)
            self.conn.execute(
                "INSERT INTO ayah_theme_groups VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (
                    record_id, source_record_id, source_row, row["theme"],
                    row["surah_number"], row["ayah_from"], row["ayah_to"],
                    row["keywords"], row["total_ayahs"],
                    hashlib.sha256(exact.encode("utf-8")).hexdigest(),
                ),
            )
            if valid:
                for ayah in range(row["ayah_from"], row["ayah_to"] + 1):
                    key = f"{row['surah_number']}:{ayah}"
                    covered.add(key)
                    self.conn.execute(
                        "INSERT INTO ayah_theme_group_ayahs VALUES (?, ?)", (record_id, key)
                    )
        duplicates = len(rows) - len(set(exact_keys))
        gaps = len(canonical_keys - covered)
        self.finding("themes", "THEME_PHYSICAL_ROW_COUNT", "error", len(rows) == 2098, len(rows), 2098)
        self.finding("themes", "THEME_UNIQUE_GROUP_COUNT", "error", len(set(exact_keys)) == 1049, len(set(exact_keys)), 1049)
        self.finding(
            "themes", "THEME_EXACT_DUPLICATES_PRESERVED", "info",
            duplicates == 1049, duplicates, 1049, "Expected source condition; duplicates remain in staging."
        )
        self.finding("themes", "THEME_RANGE_VALIDITY", "error", invalid_ranges == 0, invalid_ranges, 0)
        self.finding("themes", "THEME_TOTAL_AYAH_CONSISTENCY", "error", total_mismatches == 0, total_mismatches, 0)
        self.finding(
            "themes", "THEME_COVERAGE_GAPS_PRESERVED", "info",
            gaps == 36, gaps, 36, "Expected source condition; gaps are not synthesized."
        )
        self.metrics["themes"] = {
            "physical_rows": len(rows),
            "unique_groups": len(set(exact_keys)),
            "exact_duplicates": duplicates,
            "covered_ayahs": len(covered),
            "coverage_gaps": gaps,
        }

    def load_hadith(self) -> None:
        source_id, path = SOURCES["hadith"]
        data = json_load(path)
        metadata = data["metadata"]
        rows = data["hadiths"]
        sections = metadata["sections"]
        details = metadata["section_details"]
        numbers = set()
        blank_numbers = []
        unmapped = grade_count = 0
        for index, raw in enumerate(rows):
            number = int(raw["hadithnumber"])
            key = str(number)
            numbers.add(number)
            source_record_id = self.source_record(source_id, key, path, f"/hadiths/{index}", raw)
            section_keys = [
                section for section, detail in details.items()
                if int(detail["hadithnumber_first"]) <= number <= int(detail["hadithnumber_last"])
                and int(section) != 0
            ]
            unmapped += len(section_keys) != 1
            section_key = section_keys[0] if section_keys else None
            reference = raw.get("reference") or {}
            record_id = stable_id(source_id, key)
            self.conn.execute(
                "INSERT INTO hadith_records VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                (
                    record_id, source_record_id, "abudawud", number,
                    raw.get("arabicnumber"), reference.get("book"), reference.get("hadith"),
                    sections.get(section_key) if section_key else None,
                ),
            )
            text = str(raw.get("text", ""))
            if not text.strip():
                blank_numbers.append(number)
            self.conn.execute(
                "INSERT INTO hadith_text_versions VALUES (?, ?, ?, ?)",
                (stable_id(record_id, "en"), record_id, "en", text),
            )
            for ordinal, grade in enumerate(raw.get("grades") or [], 1):
                grade_count += 1
                self.conn.execute(
                    "INSERT INTO hadith_grade_assertions VALUES (?, ?, ?, ?)",
                    (
                        stable_id(record_id, "grade", ordinal),
                        record_id,
                        str(grade.get("name", "")),
                        str(grade.get("grade", "")),
                    ),
                )
        self.finding("hadith", "HADITH_RECORD_COUNT", "error", len(rows) == 5274, len(rows), 5274)
        self.finding("hadith", "HADITH_UNIQUE_NUMBERS", "error", len(numbers) == len(rows), len(numbers), len(rows))
        self.finding(
            "hadith",
            "HADITH_BLANK_TEXT_PRESERVED",
            "warning",
            blank_numbers == [907, 4290],
            ",".join(map(str, blank_numbers)),
            "907,4290",
            "Source rows retain identity, references, and grades; no text was synthesized.",
        )
        self.finding("hadith", "HADITH_SECTION_MAPPING", "error", unmapped == 0, unmapped, 0)
        self.finding("hadith", "HADITH_GRADE_ASSERTION_COUNT", "error", grade_count == 18818, grade_count, 18818)
        self.metrics["hadith"] = {
            "records": len(rows),
            "grade_assertions": grade_count,
            "sections": len(sections) - (1 if "0" in sections else 0),
            "blank_text_hadith_numbers": blank_numbers,
        }

    def load_verification(self) -> None:
        source_id, path = SOURCES["verification"]
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        all_rows = list(sheet.iter_rows(values_only=True))
        headers = [str(value or "") for value in all_rows[0]]
        rows = all_rows[1:]
        reference_indexes = [index for index, header in enumerate(headers) if header.strip().upper() == "RUJUKAN"]
        core_blanks = reference_count = 0
        statuses = Counter()
        for source_row, values in enumerate(rows, 2):
            raw = {f"{index + 1}:{headers[index]}": value for index, value in enumerate(values)}
            source_record_id = self.source_record(
                source_id, str(source_row), path, f"{sheet.title}!{source_row}", raw
            )
            text_ar = values[1] if len(values) > 1 else None
            text_ms = values[2] if len(values) > 2 else None
            status = values[6] if len(values) > 6 else None
            researcher = values[8] if len(values) > 8 else None
            core_blanks += sum(not str(value or "").strip() for value in (text_ar, text_ms, status))
            statuses[str(status or "").strip()] += 1
            claim_id = stable_id(source_id, source_row)
            self.conn.execute(
                "INSERT INTO hadith_verification_claims VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (
                    claim_id, source_record_id, source_row, text_ar, text_ms,
                    values[3] if len(values) > 3 else None,
                    values[4] if len(values) > 4 else None,
                    values[5] if len(values) > 5 else None,
                    status, values[7] if len(values) > 7 else None, researcher,
                ),
            )
            for ordinal, ref_index in enumerate(reference_indexes, 1):
                reference = values[ref_index] if ref_index < len(values) else None
                volume_page = values[ref_index + 1] if ref_index + 1 < len(values) else None
                if not str(reference or "").strip() and not str(volume_page or "").strip():
                    continue
                reference_count += 1
                self.conn.execute(
                    "INSERT INTO hadith_verification_references VALUES (?, ?, ?, ?, ?)",
                    (
                        stable_id(claim_id, "reference", ordinal),
                        claim_id, ordinal, reference, volume_page,
                    ),
                )
        workbook.close()
        self.finding("verification", "SEMAK_RECORD_COUNT", "error", len(rows) == 60, len(rows), 60)
        self.finding("verification", "SEMAK_CORE_REQUIRED_FIELDS", "error", core_blanks == 0, core_blanks, 0)
        self.finding(
            "verification", "SEMAK_REFERENCES_PRESERVED", "info",
            reference_count > 0, reference_count, "> 0",
            "Repeated RUJUKAN columns are retained as ordered references."
        )
        self.metrics["verification"] = {
            "claims": len(rows),
            "references": reference_count,
            "raw_status_vocabulary": dict(statuses),
        }

    def write_outputs(self) -> None:
        fields = [
            "domain", "rule_code", "severity", "result", "observed",
            "expected", "source_record_key", "notes",
        ]
        with (self.output_dir / "validation_results.csv").open(
            "w", encoding="utf-8-sig", newline=""
        ) as handle:
            writer = csv.DictWriter(handle, fieldnames=fields)
            writer.writeheader()
            writer.writerows(self.results)
        failed_errors = [
            row for row in self.results
            if row["result"] == "FAIL" and row["severity"] == "error"
        ]
        summary = {
            "import_run_id": RUN_ID,
            "status": "passed" if not failed_errors else "failed",
            "generated_at": utcnow(),
            "database": self.db_path.relative_to(ROOT).as_posix(),
            "validation_rule_count": len(self.results),
            "passed_rule_count": sum(row["result"] == "PASS" for row in self.results),
            "failed_rule_count": sum(row["result"] == "FAIL" for row in self.results),
            "failed_error_count": len(failed_errors),
            "metrics": self.metrics,
        }
        (self.output_dir / "import_summary.json").write_text(
            json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
        )
        return summary

    def run(self) -> dict[str, object]:
        self.output_dir.mkdir(parents=True, exist_ok=True)
        if self.db_path.exists():
            self.db_path.unlink()
        missing = [str(path) for _, path in SOURCES.values() if not path.exists()]
        if missing:
            raise FileNotFoundError("Missing source files:\n" + "\n".join(missing))
        self.conn = sqlite3.connect(self.db_path)
        try:
            self.create_schema()
            canonical_keys = self.load_quran()
            self.load_translations(canonical_keys)
            self.load_tafsir(canonical_keys)
            self.load_topics(canonical_keys)
            self.load_themes(canonical_keys)
            self.load_hadith()
            self.load_verification()
            failed_errors = sum(
                row["result"] == "FAIL" and row["severity"] == "error"
                for row in self.results
            )
            status = "passed" if failed_errors == 0 else "failed"
            self.conn.execute(
                "UPDATE import_runs SET completed_at = ?, status = ? WHERE import_run_id = ?",
                (utcnow(), status, RUN_ID),
            )
            self.conn.commit()
            return self.write_outputs()
        finally:
            self.conn.close()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=DEFAULT_OUTPUT,
        help="Output directory (default: data/staging_reports/day8)",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    output_dir = args.output_dir
    if not output_dir.is_absolute():
        output_dir = ROOT / output_dir
    try:
        summary = Prototype(output_dir).run()
    except Exception as exc:
        print(f"Day 8 import failed: {exc}", file=sys.stderr)
        return 1
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0 if summary["status"] == "passed" else 2


if __name__ == "__main__":
    raise SystemExit(main())
