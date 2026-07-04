from __future__ import annotations

import csv
import gzip
import hashlib
import json
import os
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable

import openpyxl
import pandas as pd
import pyarrow.parquet as pq


ROOT = Path(__file__).resolve().parents[1]
RAW = ROOT / "data" / "raw" / "hadith"
OUT = ROOT / "data" / "staging_reports" / "hadith"
OUT.mkdir(parents=True, exist_ok=True)

COLLECTION_ALIASES = {
    "bukhari": "bukhari",
    "sahih al bukhari": "bukhari",
    "sahih al-bukhari": "bukhari",
    "muslim": "muslim",
    "sahih muslim": "muslim",
    "abudawud": "abudawud",
    "abu dawud": "abudawud",
    "sunan abi dawud": "abudawud",
    "sunan abu dawood": "abudawud",
    "tirmidhi": "tirmidhi",
    "jami at tirmidhi": "tirmidhi",
    "jami` at-tirmidhi": "tirmidhi",
    "nasai": "nasai",
    "sunan an nasai": "nasai",
    "sunan an-nasa'i": "nasai",
    "sunan an nasa i": "nasai",
    "ibnmajah": "ibnmajah",
    "ibn majah": "ibnmajah",
    "sunan ibn majah": "ibnmajah",
}


def clean_name(value: Any) -> str:
    text = str(value or "").lower()
    text = re.sub(r"[^a-z0-9]+", " ", text).strip()
    return COLLECTION_ALIASES.get(text, text.replace(" ", ""))


def normalized_text(value: Any) -> str:
    text = unicodedata.normalize("NFKC", str(value or "")).lower()
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"[\W_]+", " ", text, flags=re.UNICODE)
    return " ".join(text.split())


def text_hash(value: Any) -> str:
    return hashlib.sha256(normalized_text(value).encode("utf-8")).hexdigest()


def mojibake_score(value: Any) -> int:
    text = str(value or "")
    markers = ("Ã", "Â", "Ø", "Ù", "â€", "ï·", "ðŸ")
    return sum(text.count(marker) for marker in markers)


def read_json(path: Path) -> Any:
    with path.open("r", encoding="utf-8-sig") as handle:
        return json.load(handle)


def flatten_values(value: Any, limit: int = 2000) -> list[str]:
    found: list[str] = []
    stack = [value]
    while stack and len(found) < limit:
        current = stack.pop()
        if isinstance(current, str):
            found.append(current)
        elif isinstance(current, dict):
            stack.extend(current.values())
        elif isinstance(current, list):
            stack.extend(current)
    return found


def json_record_count(value: Any) -> int:
    if isinstance(value, list):
        return len(value)
    if not isinstance(value, dict):
        return 1
    for key in ("hadiths", "hadith", "data", "records", "items"):
        candidate = value.get(key)
        if isinstance(candidate, list):
            return len(candidate)
        if isinstance(candidate, dict):
            for nested in ("data", "hadiths", "records", "items"):
                nested_value = candidate.get(nested)
                if isinstance(nested_value, list):
                    return len(nested_value)
    return len(value)


def profile_json(path: Path) -> dict[str, Any]:
    value = read_json(path)
    strings = flatten_values(value)
    return {
        "format": "json",
        "rows": json_record_count(value),
        "columns": ",".join(list(value.keys())[:30]) if isinstance(value, dict) else "",
        "mojibake_markers": sum(mojibake_score(v) for v in strings),
        "blank_strings_sampled": sum(1 for v in strings if not v.strip()),
    }


def profile_csv(path: Path) -> dict[str, Any]:
    rows = 0
    blank_cells = 0
    mojibake = 0
    columns: list[str] = []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        columns = reader.fieldnames or []
        for row in reader:
            rows += 1
            for value in row.values():
                if value is None or not str(value).strip():
                    blank_cells += 1
                else:
                    mojibake += mojibake_score(value)
    return {
        "format": "csv",
        "rows": rows,
        "columns": ",".join(columns),
        "mojibake_markers": mojibake,
        "blank_cells": blank_cells,
    }


def profile_jsonl(path: Path, compressed: bool = False) -> dict[str, Any]:
    opener = gzip.open if compressed else open
    rows = 0
    invalid = 0
    keys: Counter[str] = Counter()
    mojibake = 0
    with opener(path, "rt", encoding="utf-8-sig") as handle:
        for line in handle:
            if not line.strip():
                continue
            try:
                value = json.loads(line)
                rows += 1
                if isinstance(value, dict):
                    keys.update(value.keys())
                mojibake += sum(mojibake_score(v) for v in flatten_values(value, 100))
            except json.JSONDecodeError:
                invalid += 1
    return {
        "format": "jsonl.gz" if compressed else "jsonl",
        "rows": rows,
        "columns": ",".join(keys.keys()),
        "invalid_rows": invalid,
        "mojibake_markers": mojibake,
    }


def profile_xlsx(path: Path) -> dict[str, Any]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = []
    total_rows = 0
    for sheet in workbook.worksheets:
        rows = max(sheet.max_row - 1, 0)
        total_rows += rows
        headers = [str(cell.value or "") for cell in next(sheet.iter_rows(max_row=1))]
        sheets.append(f"{sheet.title}({rows}):{','.join(headers)}")
    workbook.close()
    return {
        "format": "xlsx",
        "rows": total_rows,
        "columns": " | ".join(sheets),
    }


def profile_parquet(path: Path) -> dict[str, Any]:
    parquet = pq.ParquetFile(path)
    metadata = parquet.metadata
    return {
        "format": "parquet",
        "rows": metadata.num_rows,
        "row_groups": metadata.num_row_groups,
        "columns": ",".join(parquet.schema_arrow.names),
        "created_by": metadata.created_by or "",
    }


def profile_sql(path: Path) -> dict[str, Any]:
    text = path.read_text(encoding="utf-8", errors="replace")
    tables = re.findall(r"CREATE TABLE\s+[`\"]?([^`\"\s(]+)", text, flags=re.I)
    inserts = len(re.findall(r"INSERT INTO", text, flags=re.I))
    return {
        "format": "sql",
        "rows": inserts,
        "columns": ",".join(sorted(set(tables))),
        "mojibake_markers": mojibake_score(text[:200000]),
    }


def source_name(path: Path) -> str:
    rel = path.relative_to(RAW)
    return rel.parts[1] if len(rel.parts) > 1 else rel.parts[0]


def selected_payloads() -> list[Path]:
    paths: list[Path] = []
    for directory, child_dirs, filenames in os.walk(RAW):
        current = Path(directory)
        rel_dir = current.relative_to(RAW).as_posix()
        if rel_dir.endswith("fawaz-hadith-api-v1/editions"):
            child_dirs.clear()
        child_dirs[:] = [
            name
            for name in child_dirs
            if name != ".git"
            and not (
                "fawaz-hadith-api-v1" in rel_dir
                and (name == "database" or "/database/" in f"{rel_dir}/{name}/")
            )
        ]
        if "/by_chapter" in rel_dir or "/min" in rel_dir:
            child_dirs.clear()
            continue
        for filename in filenames:
            path = current / filename
            suffix = path.suffix.lower()
            rel = path.relative_to(RAW).as_posix()
            if suffix in {".csv", ".xlsx", ".parquet", ".sql", ".jsonl", ".gz"}:
                paths.append(path)
            elif suffix == ".json":
                if "fawaz-hadith-api-v1/editions/" in rel and ".min.json" not in rel:
                    paths.append(path)
                elif "/by_chapter/" in rel or "/min/" in rel:
                    continue
                elif path.stat().st_size >= 100_000 or source_name(path) in {
                    "semakhadis-api",
                    "semakhadis-frontend",
                    "ahmedbaset-hadith-api",
                }:
                    paths.append(path)
    return sorted(set(paths))


def profile_payloads() -> list[dict[str, Any]]:
    results = []
    payloads = selected_payloads()
    print(f"Profiling {len(payloads)} principal payloads", flush=True)
    for index, path in enumerate(payloads, start=1):
        if index == 1 or index % 25 == 0:
            print(
                f"[{index}/{len(payloads)}] {path.relative_to(ROOT).as_posix()}",
                flush=True,
            )
        row: dict[str, Any] = {
            "source": source_name(path),
            "path": path.relative_to(ROOT).as_posix(),
            "bytes": path.stat().st_size,
            "status": "pass",
            "error": "",
        }
        try:
            suffix = path.suffix.lower()
            if suffix == ".json":
                row.update(profile_json(path))
            elif suffix == ".csv":
                row.update(profile_csv(path))
            elif suffix == ".jsonl":
                row.update(profile_jsonl(path))
            elif suffix == ".gz":
                row.update(profile_jsonl(path, compressed=True))
            elif suffix == ".xlsx":
                row.update(profile_xlsx(path))
            elif suffix == ".parquet":
                row.update(profile_parquet(path))
            elif suffix == ".sql":
                row.update(profile_sql(path))
        except Exception as exc:
            row["status"] = "fail"
            row["error"] = f"{type(exc).__name__}: {exc}"
        results.append(row)
        if index % 100 == 0:
            print(f"profiled {index} payloads", flush=True)
    return results


def extract_meeatif() -> dict[str, list[dict[str, Any]]]:
    root = RAW / "multilingual" / "hf-meeatif-hadith-datasets"
    collections: dict[str, list[dict[str, Any]]] = {}
    for path in root.glob("*.json"):
        value = read_json(path)
        if not isinstance(value, list):
            continue
        name = clean_name(path.stem)
        rows = []
        for item in value:
            rows.append(
                {
                    "number": str(item.get("In-book reference") or ""),
                    "arabic": item.get("Arabic_Text") or "",
                    "english": item.get("English_Text") or "",
                    "grade": item.get("Grade") or "",
                }
            )
        collections[name] = rows
    return collections


def extract_ahmedbaset() -> dict[str, list[dict[str, Any]]]:
    root = (
        RAW
        / "quarantined"
        / "ahmedbaset-hadith-json-v1.2.0"
        / "db"
        / "by_book"
        / "the_9_books"
    )
    collections: dict[str, list[dict[str, Any]]] = {}
    for path in root.glob("*.json"):
        value = read_json(path)
        rows_value = value.get("hadiths", []) if isinstance(value, dict) else []
        rows = []
        for item in rows_value:
            english = item.get("english") or {}
            rows.append(
                {
                    "number": str(item.get("idInBook") or item.get("id") or ""),
                    "arabic": item.get("arabic") or "",
                    "english": english.get("text") or "",
                    "grade": "",
                }
            )
        collections[clean_name(path.stem)] = rows
    return collections


def extract_fawaz() -> dict[str, list[dict[str, Any]]]:
    root = RAW / "collections" / "fawaz-hadith-api-v1" / "editions"
    collections: dict[str, dict[str, list[dict[str, Any]]]] = defaultdict(dict)
    for path in root.glob("*.json"):
        if path.name.endswith(".min.json"):
            continue
        match = re.match(r"(ara|eng)-(.+?)(?:1)?\.json$", path.name)
        if not match:
            continue
        language, book = match.groups()
        language = {"ara": "arabic", "eng": "english"}[language]
        if path.stem.endswith("1"):
            continue
        value = read_json(path)
        rows_value = value.get("hadiths", []) if isinstance(value, dict) else []
        rows = []
        for item in rows_value:
            rows.append(
                {
                    "number": str(item.get("hadithnumber") or item.get("hadithNumber") or ""),
                    language: item.get("text") or "",
                    "grade": item.get("grades") or [],
                }
            )
        collections[clean_name(book)][language] = rows

    merged: dict[str, list[dict[str, Any]]] = {}
    for book, languages in collections.items():
        by_number: dict[str, dict[str, Any]] = {}
        for language, rows in languages.items():
            for item in rows:
                number = item["number"]
                row = by_number.setdefault(
                    number, {"number": number, "arabic": "", "english": "", "grade": ""}
                )
                row[language] = item.get(language, "")
                if item.get("grade"):
                    row["grade"] = json.dumps(item["grade"], ensure_ascii=False)
        merged[book] = list(by_number.values())
    return merged


def extract_lk() -> dict[str, list[dict[str, Any]]]:
    root = RAW / "research" / "lk-hadith-corpus"
    folder_map = {
        "Bukhari": "bukhari",
        "Muslim": "muslim",
        "AbuDaud": "abudawud",
        "Tirmidhi": "tirmidhi",
        "Tirmizi": "tirmidhi",
        "Nasai": "nasai",
        "Nesai": "nasai",
        "IbnMaja": "ibnmajah",
        "IbnMajah": "ibnmajah",
    }
    collections: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for folder, name in folder_map.items():
        directory = root / folder
        if not directory.exists():
            continue
        for path in sorted(directory.glob("*.csv")):
            with path.open("r", encoding="utf-8-sig", newline="") as handle:
                reader = csv.DictReader(handle)
                for item in reader:
                    collections[name].append(
                        {
                            "number": str(item.get("Hadith_number") or ""),
                            "arabic": item.get("Arabic_Hadith") or "",
                            "english": item.get("English_Hadith") or "",
                            "grade": item.get("English_Grade") or item.get("Arabic_Grade") or "",
                        }
                    )
    return dict(collections)


def extract_abdullah() -> dict[str, list[dict[str, Any]]]:
    root = RAW / "collections" / "abdullah-naseer-six-books" / "src"
    name_map = {
        "sahih-bukhari": "bukhari",
        "sahih-muslim": "muslim",
        "abu-dawood": "abudawud",
        "al-tirmidhi": "tirmidhi",
        "sunan-nasai": "nasai",
        "ibn-e-majah": "ibnmajah",
    }
    collections: dict[str, list[dict[str, Any]]] = {}
    for path in root.glob("*.json"):
        value = read_json(path)
        hadiths = value.get("hadiths", {}) if isinstance(value, dict) else {}
        rows_value: Any = hadiths
        if isinstance(hadiths, dict):
            for key in ("data", "hadiths", "items"):
                if isinstance(hadiths.get(key), list):
                    rows_value = hadiths[key]
                    break
        if not isinstance(rows_value, list):
            rows_value = []
        rows = []
        for item in rows_value:
            rows.append(
                {
                    "number": str(
                        item.get("hadithNumber")
                        or item.get("hadith_number")
                        or item.get("idInBook")
                        or item.get("id")
                        or ""
                    ),
                    "arabic": (
                        item.get("arabic")
                        or item.get("arabicText")
                        or item.get("hadithArabic")
                        or ""
                    ),
                    "english": (
                        item.get("english")
                        or item.get("englishText")
                        or item.get("hadithEnglish")
                        or ""
                    ),
                    "grade": item.get("grade") or item.get("status") or "",
                }
            )
        collections[name_map.get(path.stem, clean_name(path.stem))] = rows
    return collections


def source_metrics(
    extracted: dict[str, dict[str, list[dict[str, Any]]]]
) -> list[dict[str, Any]]:
    rows = []
    for source, collections in extracted.items():
        for collection, records in collections.items():
            arabic_hashes = [text_hash(r["arabic"]) for r in records if normalized_text(r["arabic"])]
            english_hashes = [
                text_hash(r["english"]) for r in records if normalized_text(r["english"])
            ]
            numbers = [r["number"] for r in records if r["number"]]
            rows.append(
                {
                    "source": source,
                    "collection": collection,
                    "records": len(records),
                    "arabic_nonblank": len(arabic_hashes),
                    "english_nonblank": len(english_hashes),
                    "grade_nonblank": sum(1 for r in records if str(r["grade"]).strip()),
                    "number_nonblank": len(numbers),
                    "duplicate_numbers": len(numbers) - len(set(numbers)),
                    "duplicate_arabic_hashes": len(arabic_hashes) - len(set(arabic_hashes)),
                    "duplicate_english_hashes": len(english_hashes) - len(set(english_hashes)),
                    "mojibake_markers": sum(
                        mojibake_score(r["arabic"]) + mojibake_score(r["english"])
                        for r in records
                    ),
                }
            )
    return rows


def compare_sources(
    extracted: dict[str, dict[str, list[dict[str, Any]]]]
) -> list[dict[str, Any]]:
    source_names = sorted(extracted)
    rows = []
    for i, left_name in enumerate(source_names):
        for right_name in source_names[i + 1 :]:
            common_collections = sorted(
                set(extracted[left_name]).intersection(extracted[right_name])
            )
            for collection in common_collections:
                left = extracted[left_name][collection]
                right = extracted[right_name][collection]
                for language in ("arabic", "english"):
                    left_hashes = {
                        text_hash(row[language])
                        for row in left
                        if normalized_text(row[language])
                    }
                    right_hashes = {
                        text_hash(row[language])
                        for row in right
                        if normalized_text(row[language])
                    }
                    overlap = left_hashes.intersection(right_hashes)
                    union = left_hashes.union(right_hashes)
                    rows.append(
                        {
                            "left_source": left_name,
                            "right_source": right_name,
                            "collection": collection,
                            "language": language,
                            "left_unique_texts": len(left_hashes),
                            "right_unique_texts": len(right_hashes),
                            "exact_normalized_overlap": len(overlap),
                            "jaccard": round(len(overlap) / len(union), 6) if union else 0,
                        }
                    )
    return rows


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    keys = sorted({key for row in rows for key in row})
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=keys)
        writer.writeheader()
        writer.writerows(rows)


def main() -> int:
    profiles = profile_payloads()
    write_csv(OUT / "payload_validation.csv", profiles)

    extracted = {
        "meeatif": extract_meeatif(),
        "ahmedbaset": extract_ahmedbaset(),
        "fawaz": extract_fawaz(),
        "lk": extract_lk(),
        "abdullah_naseer": extract_abdullah(),
    }
    metrics = source_metrics(extracted)
    comparisons = compare_sources(extracted)
    write_csv(OUT / "collection_metrics.csv", metrics)
    write_csv(OUT / "cross_source_comparison.csv", comparisons)

    summary = {
        "payload_files_profiled": len(profiles),
        "payload_pass": sum(1 for row in profiles if row["status"] == "pass"),
        "payload_fail": sum(1 for row in profiles if row["status"] == "fail"),
        "parquet_files": sum(1 for row in profiles if row.get("format") == "parquet"),
        "parquet_rows": sum(
            int(row.get("rows") or 0)
            for row in profiles
            if row.get("format") == "parquet" and row["status"] == "pass"
        ),
        "source_collection_profiles": len(metrics),
        "comparison_rows": len(comparisons),
        "sources": sorted(extracted),
        "failed_payloads": [
            {"path": row["path"], "error": row["error"]}
            for row in profiles
            if row["status"] == "fail"
        ],
    }
    (OUT / "validation_summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(json.dumps(summary, indent=2))
    return 1 if summary["payload_fail"] else 0


if __name__ == "__main__":
    raise SystemExit(main())
